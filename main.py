import csv
import numpy as np
import pandas as pd
import os.path
import openpyxl
import gc
import time
import datetime
import argparse
from collections import namedtuple
import math
from lib.tensor_sketch import TS, MH
from lib.tensor_sketch import TE
from lib.vector_sketch import RefIdx
from lib.base import SketchParams

Vectorizer = namedtuple('Vectorizer', ['func', 'name'])
params = SketchParams(A=4, t=6, D=128)
ts = TS(params)
tensor_sketch = Vectorizer(func=ts.sketch, name = 'TS')
    
#distance, true_distance, correct, is_in_pool, read_offset, matched_offset, true_offset   
def classify_csv(filepath, th=math.inf):
    with open(filepath,'r') as f:

        reader = csv.reader(f)
        l = [row[:4] for row in reader]
        data = [[float(el) for el in row] for row in l[1:]]
        
        query_num = len(data)
        tp = 0
        fp = 0 
        tn = 0
        fn = 0
        missr = 0
        miss_num = 0
        for row in data:
            
            dist = row[0]
            
            correct = int(row[2])
            is_in_pool = bool(int(row[3]))

            if is_in_pool and not correct:
                true_dist = row[1]
                miss_num += 1
                if true_dist < dist:
                    missr += 1

            if (dist<=th):
                if is_in_pool:
                    tp += correct
                    fp += 1-correct
                else:
                    fn += 1

            else:
                if(is_in_pool):
                    fp +=1
                else:
                    tn += 1

    fpr = -1 if tp+fp==0 else fp/(tp+fp)
    fnr = -1 if tn+fn==0 else fn/(tn+fn)
    missr = 0 if miss_num == 0 else missr/miss_num
    return query_num, fpr, fnr, missr



def build_Vec(filename, vectorizer=tensor_sketch, indextype = 'annoy', w: int = 120, o:int = 100, sketch_dim: int = 16, rebuild=True, wf=False):
    
    idx = RefIdx(filename = filename, vectorizer = vectorizer,sketchdim = sketch_dim, w=w, o=o, index=indextype, rebuild=rebuild,  wf=wf)
    
    return idx, idx.vectorizing_time, idx.build_time

def build_TE(filename, subseq_len: int = 3, indextype = 'annoy', w: int = 120, o:int = 100,tensor_sketch_dim: int = 16, rebuild=True, wf=False):
    params = SketchParams(A=4, t=subseq_len, D=tensor_sketch_dim)
    te = TE(params)
    
    tensor_embedding = Vectorizer(func=te.sketch, name = 'TE')
    start_time = time.time()
    idx = RefIdx(filename = filename, vectorizer = tensor_embedding, w=w, o=o, index=indextype, rebuild=rebuild,  wf=wf)
    total_time = time.time() - start_time
    return idx, total_time

def build_TS(filename, subseq_len: int = 3, indextype = 'annoy', w: int = 120, o:int = 100,tensor_sketch_dim: int = 16, rebuild=True, wf=False):
    params = SketchParams(A=4, t=subseq_len, D=tensor_sketch_dim)
    ts = TS(params)
    
    tensor_sketching = Vectorizer(func=ts.sketch, name = 'TS')
    
    idx = RefIdx(filename = filename, vectorizer = tensor_sketching,sketchdim = tensor_sketch_dim, w=w, o=o, index=indextype, rebuild=rebuild,  wf=wf)
    
    return idx, idx.vectorizing_time, idx.build_time

def build_MH(filename, subseq_len: int = 3, indextype = 'annoy', w: int = 120, o:int = 100,tensor_sketch_dim: int = 16, rebuild=True, wf=False):
    params = SketchParams(A=4, t=subseq_len, D=tensor_sketch_dim)
    mh = MH(params)
    
    tensor_sketching = Vectorizer(func=mh.mh_sketch, name='MH')
    start_time = time.time()
    idx = RefIdx(filename = filename, vectorizer = tensor_sketching,sketchdim = tensor_sketch_dim, w=w, o=o, index=indextype, rebuild=rebuild,  wf=wf)
    total_time = time.time() - start_time
    return idx, total_time

def eval(filename, fread_file, subseq_len, indextype, w_list, s_list, rebuild= True, wf=False, sketch_dim = 16, th=math.inf):

    outfile = '/cluster/scratch/junbzhang/out/ts_{}_{}.csv'.format(indextype, subseq_len)
    res = np.zeros((len(w_list),len(s_list)))
    res_fnr = np.zeros((len(w_list),len(s_list)))
    build_table = [ ['']*len(s_list)for _ in range(len(w_list))]
    query_table = [ ['']*len(s_list)for _ in range(len(w_list))]
    for i,w in enumerate(w_list):
        for j,s in enumerate(s_list):
            idx, build_time = build_TS(filename, subseq_len = subseq_len, indextype = indextype, w = w, o = w-s, rebuild=rebuild, wf=wf, tensor_sketch_dim = sketch_dim)
            start_time = time.time()
            idx.query_file(fread_file, outfile, check_correct=True, frac=1.0, ws=False)
            query_time = time.time() - start_time
            del idx
            build_table[i][j] = str(datetime.timedelta(seconds=round(build_time)))
            query_table[i][j] = str(datetime.timedelta(seconds=round(query_time)))
            gc.collect()
            #query_num, uniq_r, fpr, match_r , neigh_r= get_score_csv(outfile)
            _, fpr, fnr = classify_csv(outfile,th)
            res[i][j] = fpr
            res_fnr[i][j] = fnr
            print('finished', subseq_len, w, s)
    
    return res, build_table, query_table, res_fnr

def table_to_csv(table, file_path,file_name):
    filepath = '{}/{}.csv'.format(file_path, file_name)
    with open(filepath,'w') as f:
        
        csvWriter = csv.writer(f,delimiter=',')
        csvWriter.writerows(table)

    return 

def write_to_csv(data, file_path, file_name):
    filepath = '{}/{}.csv'.format(file_path, file_name)
    with open(filepath,'a+') as f:
        f.write(data)
    return 

def table_to_excel(res, file_path, file_name, sheet_name, subseq_len, row = 0, col = 0, index = [], columns = []):
        filepath = '{}/{}.xlsx'.format(file_path, file_name)
        first_time = False
        
        if not os.path.isfile(filepath):
            df_empty = pd.DataFrame()
            df_empty.to_excel(filepath)
            first_time = True
            
        workbook = openpyxl.load_workbook(filepath)
        
        if first_time and not (sheet_name in workbook.sheetnames):
            ex_sheet = workbook['Sheet1']
            ex_sheet.title = sheet_name
            workbook.save(filepath)
        workbook.close()
        
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer: 
            df = pd.DataFrame(res, index = index, columns = columns)
            df.to_excel(writer, sheet_name=sheet_name, startrow = row, startcol = col)
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook[sheet_name]
        worksheet.cell(row=row+1, column = col+1).value = 't = {}'.format(subseq_len)
        workbook.save(filepath)
        workbook.close()

class FloatRange(object):
            def __init__(self, start: float, end: float):
                self.start = start
                self.end = end

            def __eq__(self, other: float) -> bool:
                return self.start <= other <= self.end

            def __repr__(self):
                return '[{0},{1}]'.format(self.start, self.end)

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-FR', '--file_path_ref', default = 'scratch/data')
    parser.add_argument('-R', '--reference_file', default = 'filter_random_ref.fasta')
    parser.add_argument('-FQ', '--file_path_query', default = 'scratch/data')
    parser.add_argument('-Q', '--query_file', default = 'filter_random_read.fasta')
    parser.add_argument('-V', '--vectorizer', choices=['kmer-dist', 'kmer-pos', 'tensor_embedding','min_hash'], default='tensor_sketch')
    parser.add_argument('-I', '--index_type', choices=['faiss', 'annoy'] ,default = 'annoy')
    parser.add_argument('-K', '--kmer_length', type=int, default=6)  
    parser.add_argument('-D', '--sketch_dim', type=int, default=128)
    parser.add_argument('-W', '--window_size', type=int, choices=range(10, 180), default=100)
    parser.add_argument('-O', '--overlap_frac', type=float, choices=[FloatRange(0.01, 0.3)], default=.1)
    parser.add_argument('-S', '--stride', type=int, default=1)
    parser.add_argument('-F', '--query_frac', type=float, choices=[FloatRange(0.01, 1.0)], default=1.0)
    parser.add_argument('-P', '--out_prefix', default='scratch/eval')
    parser.add_argument('-T', '--tmp_prefix', default='scratch/out')
    parser.add_argument('-B', '--rebuild_index', action='store_false')
    args = parser.parse_args()
    #subseq_len, indextype, w, o, tensor_sketch_dim
    reference_file = args.file_path_ref + '/' + args.reference_file
    query_file = args.file_path_query + '/' + args.query_file
    vectorizer = args.vectorizer
    index_type = args.index_type
    sketch_dim = args.sketch_dim
    kmer_len = args.kmer_length
    window = args.window_size
    overlap_frac = args.overlap_frac
    stride = args.stride
    out_prefix = args.out_prefix
    tmp_prefix = args.tmp_prefix
    rebuild = args.rebuild_index

    excel_name = '{}_{}'.format(index_type,vectorizer)
    sheetname = '{}_{}'.format(kmer_len,sketch_dim)
    offset = 0


    
    out_file  =  '{}/{}_{}_({},{},{},{}).csv'.format(tmp_prefix,index_type, vectorizer, sketch_dim, kmer_len,window,stride)

    idx, vectorizing_time, build_time = build_TS(reference_file, subseq_len = kmer_len, indextype = index_type, w = window, o = window-stride, rebuild=rebuild, wf=False, tensor_sketch_dim = sketch_dim)

    query_time = idx.query_file(query_file, out_file, check_correct=True, frac=1.0, ws=False)
    query_num, fpr, fnr, missr = classify_csv(out_file)

    def format_time(num):
        tmp = str(datetime.timedelta(seconds=round(num,2))).split('.')
        print(num)
        return tmp[0]+ '.' + tmp[1][:2]

    vectorizing_time = round(vectorizing_time, 2)
    build_time = round(build_time, 2)
    query_time = round(query_time, 2)
    str_row = '{},{},{},{},{},{},{},{},{}\n'.format(sketch_dim,kmer_len, window, stride, fpr, missr, vectorizing_time, build_time, query_time)
    csv_name = '{}_{}_param_analysis'.format(index_type, vectorizer)
    write_to_csv(str_row, out_prefix, csv_name)
    
    
    