import lib.vector_sketch as vs
from lib.tebns import TE
from lib.base import *
from collections import namedtuple
import pandas as pd
import openpyxl
from annoy import AnnoyIndex
import numba as nb
from lib.tensor_sketch import TS
from lib.tensor_sketch import MH
from hashlib import sha1
import math
import Levenshtein
def test3():
    seq = 'AAATGGAGCCGGAGACCGGCGTCGACGCGGTGAACGGGTCCCGCCGAACCCCTCGAAGCAGCCGCCGCGAAGGTGGCCGCCACGATCCAGGACAGCAGGCGGGCG'
    print(len(seq))
    es = vs.encode_seq(seq)
    kmers = vs.seq2kmers(es)
    print(es)
    print(kmers)

def testTS():
    
                    
    matched_ref = 'GACCGCGAGCAAATCGAGGATCACGTTGCGAAGCGATTGACCGCCGAACGTGCCCAGTTG'
                  
    true_ref2 =   'GACCACCGGGAGCACAAGTGCTGACGAACGTTGTGATCGACCAGAGCCGGACCTTCATGG'
                  
    

    true_ref =    'CCCAGACCACCAGGAGCACAAGATGCTGACGAACGTTGTGATCGACCAGAGCCGGACCTT'
                      
                      
    read_seq =    'GACCACCGGGAGCACAAGTGCTGACGAACGTTGTGATCGACCAGAGCCGGACCTTCATGG'
    tensor_sketch_dim = 256
    te_subseq_len =11
    
    params = SketchParams(A=4, t=te_subseq_len, D=tensor_sketch_dim)
    ts = TS(params)
    sk_true = ts.sketch(true_ref2)
    sk_match = ts.sketch(matched_ref)
    sk_read = ts.sketch(read_seq)
    print(math.dist(sk_true,sk_read))
    print(math.dist(sk_match,sk_read))
    print(Levenshtein.distance(true_ref2,read_seq))
    print(Levenshtein.distance(matched_ref, read_seq))
testTS()
def test4():
    
    seq = 'ACTGTGACGCT'
    
    tensor_sketch_dim = 16  
    te_subseq_len = 4
    tensor_embedding_dim = 4 ** te_subseq_len
    Vectorizer = namedtuple('Vectorizer', ['func', 'ndim'])

    params = SketchParams(A=4, t=te_subseq_len, D=tensor_sketch_dim)
    te = TE(params)
    
    fe=te._full_sketch(encode_seq(seq))
    print(fe)
    te._normalize(len(seq),fe[te_subseq_len])
    arr=np.array([x for x in fe[te_subseq_len]])
    count = 0
    for el in arr:
        print(el)
        count += el
        print(0<=el<=1)
    print(count)
    fe=te.sketch(seq)
    print(fe)
    #tensor_embedding = Vectorizer(func=te.sketch, ndim=tensor_embedding_dim)
    #print(vs.seq2vec(seq, tensor_embedding))

def test5():
    matches = []
    def rettuple():
        return 0,1,0
    matches.append((rettuple()+(2)))
    print(matches)

def test6():
    indextype = 'annoy'
    w_list = [1,2,3]
    s_list = [2,3,4]
    outputfiles = [['scratch/out/te_{}_{}_{}'.format(indextype, w, s) for w in w_list]  for s in s_list]
    print(outputfiles[1][1])

def test7():
    s_list = [5, 10, 15]
    w_list = [30, 50, 80, 100]
    res = np.zeros((len(w_list),len(s_list)))
    df = pd.DataFrame(res, index = w_list, columns=s_list)
    # create new excel file
    df_empty = pd.DataFrame()
    df_empty.to_excel('scratch/eval/test.xlsx')

    # ... with a sheet name
    workbook = openpyxl.load_workbook('scratch/eval/test.xlsx')
    ex_sheet = workbook['Sheet1']
    ex_sheet.title = 'Tmp'
    workbook.save('scratch/eval/test.xlsx')

    # prepare a dataframe
    df = pd.DataFrame({'Column1': ['aa', 'bb', 'cc', 'dd'],
                    'Column2': [100, 170, 140, 160]})

    # insert a dataframe into an excel sheet
    with pd.ExcelWriter('scratch/eval/test.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        writer.workbook = openpyxl.load_workbook('scratch/eval/test.xlsx')
        df.to_excel(writer, sheet_name='Tmp', index=[], header=True, startrow=3, startcol=3)
    with pd.ExcelWriter('scratch/eval/test.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        writer.workbook = openpyxl.load_workbook('scratch/eval/test.xlsx')
        df.to_excel(writer, sheet_name='Tmp', index=False, header=True, startrow=5, startcol=7) 

def test8():
    vectors=[[1,1,1,1],[1,0,0,0], [0.1,0.2,0.3,0.4], [2,2,2,2], [3,3,3,3],[4,4,4,4]]
    for i, v in enumerate(vectors):
        t = AnnoyIndex(4, "euclidean")
        t.add_item(i, v)
    t.build(n_trees=4)
    for i, v in enumerate(vectors):
        
        ids, distances = t.get_nns_by_vector(vectors[i],1, include_distances=True)
        print(ids[0],distances[0])
        print(math.dist(vectors[ids[0]],vectors[i]))

def empty_tensor(A,t):
        pow = np.zeros(t + 1, np.int32)
        pow[0] = 1
        for i in range(1, t + 1):
            pow[i] = A * pow[i - 1]
        Ts = list()
        for l in pow:
            Ts.append(np.zeros(l, np.float64))
        return Ts

def test9(A,t):
    seq = 'AAACG'
    seq = vs.encode_seq(seq)
    Ts = empty_tensor(A,t)

    Ts[0][0] = 1

    # sketch
    loop_count = 0
    for c in seq:
        assert 0 <= c < A
        for i in range(t - 1, -1, -1):
            for j in range(len(Ts[i])):
                
                Ts[i + 1][A * j + c] += Ts[i][j]
                
                loop_count += 1
    print(loop_count)
    return Ts

def normalize(t, seq_len, full_sketch):
    # Normalization factor.
    n = seq_len
    nct = nb.float64(1)
    for i in range(t):
        nct = nct * (n - i) / (i + 1)
    full_sketch[t] /= nct

def test10(A,t):
    seq = 'AAACG'
    seq = vs.encode_seq(seq)
    Ts = empty_tensor(A,t)

    Ts[0][0] = 1
    loop_count = 0 
    for c in seq:     
        for i in range(0, min(t,loop_count+1)):
            for j in range(len(Ts[i])):
                    Ts[i+1][A * j + c] += Ts[i][j]
                    loop_count += 1
                
    print(loop_count)
    return Ts

def get_hash_func(n:int = 4): #result in bytes
        h1 = sha1()
        h2 = sha1()
        h1.update(b'2')
        h2.update(b'2')
        print(h1.digest())
        print(h2.digest())

